# 代码生成时间: 2025-08-01 08:35:44
import scrapy
from scrapy.crawler import CrawlerProcess
from scrapy.exceptions import DropItem
from openpyxl import Workbook
import logging


# 配置日志记录
logging.basicConfig(format='%(asctime)s %(levelname)s: %(message)s', level=logging.INFO)


class ExcelGeneratorSpider(scrapy.Spider):
    name = "excel_generator"
    allowed_domains = []
    start_urls = []

    def __init__(self):
        # 初始化Workbook和sheet
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.fields = []  # 定义字段
        self.data = []  # 存储数据
        self.header = []  # 存储表头

    def start_requests(self):
        # 这里可以添加模拟的start_urls，实际应用中应从外部获取
        # yield scrapy.Request(url=self.start_urls[0], callback=self.parse)
        pass

    def parse(self, response):
        # 解析响应数据，实际应用中应根据实际情况解析
        # 假设我们已经从response中解析出了数据
        self.data.append({"name": "John", "age": 30, "city": "New York"})
        self.data.append({"name": "Jane", "age": 25, "city": "Los Angeles"})

        # 处理数据
        self.process_data()

    def process_data(self):
        try:
            # 写入表头
            for index, field in enumerate(self.fields):
                self.sheet.cell(row=1, column=index + 1).value = field

            # 写入数据
            for row_index, item in enumerate(self.data, start=2):
                for col_index, field in enumerate(self.fields):
                    cell_value = item.get(field)
                    self.sheet.cell(row=row_index, column=col_index + 1).value = cell_value

            # 保存Excel文件
            self.workbook.save('output.xlsx')
            logging.info('Excel file saved successfully.')
        except Exception as e:
            logging.error(f'Error occurred while processing data: {e}')
            raise DropItem(f'Item dropped due to error: {e}')

    def close_spider(self, reason):
        # 关闭爬虫时清理资源
        logging.info(f'Spider closed due to {reason}')


# 运行爬虫
if __name__ == '__main__':
    process = CrawlerProcess(settings={"FEEDS": {"output.xlsx": {"format": "xlsx"}}})
    process.crawl(ExcelGeneratorSpider)
    process.start()